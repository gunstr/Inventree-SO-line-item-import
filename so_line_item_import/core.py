"""Custom parsing and importing of SO line items.

See docs/implementation.md for the full design rationale, in particular why
the preview/import flow re-parses the uploaded file on every request instead
of caching resolved rows server-side.
"""

from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Max
from django.db.models import Q
from django.http import JsonResponse
from django.urls import path, reverse

from order.models import SalesOrder, SalesOrderLineItem
from part.models import Part
from plugin import InvenTreePlugin
from plugin.mixins import AppMixin, UrlsMixin, UserInterfaceMixin

from . import PLUGIN_VERSION


class SOLineItemImport(AppMixin, UrlsMixin, UserInterfaceMixin, InvenTreePlugin):
    """SOLineItemImport - custom InvenTree plugin."""

    # Plugin metadata
    TITLE = "SO Line Item Import"
    NAME = "SOLineItemImport"
    SLUG = "so-line-item-import"
    DESCRIPTION = "Custom parsing and importing of SO line items"
    VERSION = PLUGIN_VERSION

    # Additional project information
    AUTHOR = "gunstr"
    WEBSITE = "https://my-project-url.com"
    LICENSE = "MIT"

    # Optionally specify supported InvenTree versions
    # MIN_VERSION = '0.18.0'
    # MAX_VERSION = '2.0.0'

    def setup_urls(self):
        """Register plugin URL endpoints."""
        return [
            path(
                "import/so-lines/",
                self.import_sales_order_lines,
                name="import-so-lines",
            )
        ]

    def _parse_excel_rows(self, upload):
        """Parse workbook rows and yield row data entries."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise RuntimeError("openpyxl is required to import Excel files")

        workbook = load_workbook(upload, data_only=True, read_only=True)
        worksheet = workbook.active

        if worksheet is None:
            return []

        header_row = None
        header_index = 0

        for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if any(cell not in [None, ""] for cell in row):
                header_row = row
                header_index = idx
                break

        if header_row is None:
            return []

        def norm(value):
            return str(value or "").strip().lower()

        def normalize_product_cell(value):
            """Normalize product cell content to lookup-friendly text.

            Example:
            "EX Extrusion:MT-EX-06-06-120-51" -> "MT-EX-06-06-120-51"
            """
            text = "" if value is None else str(value).strip()

            if not text:
                return ""

            if ":" in text:
                # Use the right-most segment so prefixed labels are ignored.
                suffix = text.rsplit(":", 1)[1].strip()
                if suffix:
                    return suffix

            return text

        normalized = [norm(cell) for cell in header_row]

        name_aliases = {
            "product",
            "product/service",
            "product / service",
            "product name",
            "part",
            "part name",
            "name",
        }
        qty_aliases = {"qty", "quantity", "order qty", "ordered quantity"}

        name_col = next(
            (i for i, col in enumerate(normalized) if col in name_aliases), None
        )
        qty_col = next(
            (i for i, col in enumerate(normalized) if col in qty_aliases), None
        )

        if name_col is None:
            name_col = 0

        if qty_col is None:
            qty_col = 1

        entries = []

        for idx, row in enumerate(
            worksheet.iter_rows(min_row=header_index + 1, values_only=True),
            start=header_index + 1,
        ):
            raw_product_name = row[name_col] if len(row) > name_col else ""
            product_name = normalize_product_cell(raw_product_name)
            quantity_value = row[qty_col] if len(row) > qty_col else None

            if not product_name and quantity_value in [None, ""]:
                continue

            entries.append({
                "row": idx,
                "product_name": product_name,
                "quantity": quantity_value,
            })

        return entries

    def _find_part_for_name(self, product_name: str):
        """Resolve a salable part from a part code or part name value."""
        value = (product_name or "").strip()

        if not value:
            return None, {
                "reason": "missing_product_name",
                "candidates": [],
            }

        # Prefer exact IPN match first, then exact name match.
        part = Part.objects.filter(IPN__iexact=value, salable=True).first()

        if not part:
            part = Part.objects.filter(name__iexact=value, salable=True).first()

        # If part exists but is not salable, report a clearer reason.
        if not part:
            non_salable_match = Part.objects.filter(
                Q(IPN__iexact=value) | Q(name__iexact=value)
            ).first()

            if non_salable_match and not non_salable_match.salable:
                return None, {
                    "reason": "part_not_salable",
                    "candidates": [],
                }

        if part:
            return part, None

        return None, {
            "reason": "part_not_found",
            "candidates": [],
        }

    def import_sales_order_lines(self, request, *args, **kwargs):
        """Import sales order line items from an uploaded Excel file.

        Handles both preview (``dry_run=true``) and the real import
        (``dry_run=false``) requests. The uploaded file is required and fully
        re-parsed/re-validated on every call - there is no server-side
        preview cache, so a dry run and the confirming import are two
        independent, consistent passes over the same data (see
        docs/implementation.md).
        """
        if request.method != "POST":
            return JsonResponse({"detail": "Method not allowed"}, status=405)

        if not request.user.is_authenticated:
            return JsonResponse({"detail": "Authentication required"}, status=401)

        can_import = request.user.has_perm("order.add_salesorderlineitem")

        if not can_import:
            return JsonResponse({"detail": "Permission denied"}, status=403)

        sales_order_id = request.POST.get("sales_order_id", None)

        if not sales_order_id:
            return JsonResponse({"detail": "Missing sales_order_id"}, status=400)

        try:
            sales_order = SalesOrder.objects.get(pk=sales_order_id)
        except (SalesOrder.DoesNotExist, ValueError):
            return JsonResponse({"detail": "Sales order not found"}, status=404)

        upload = request.FILES.get("file", None)

        dry_run = str(request.POST.get("dry_run", "false")).strip().lower() in {
            "1",
            "true",
            "yes",
            "on",
        }

        if upload is None:
            return JsonResponse({"detail": "No file uploaded"}, status=400)

        try:
            entries = self._parse_excel_rows(upload)
        except RuntimeError as exc:
            return JsonResponse({"detail": str(exc)}, status=500)
        except Exception as exc:
            return JsonResponse({"detail": f"Failed to parse file: {exc}"}, status=400)

        if not entries:
            return JsonResponse(
                {
                    "dry_run": dry_run,
                    "created_count": 0,
                    "would_create_count": 0,
                    "skipped_count": 0,
                    "errors": [],
                    "unresolved": [],
                    "preview_rows": [],
                },
                status=200,
            )

        created_count = 0
        would_create_count = 0
        skipped_count = 0
        errors = []
        unresolved = []
        preview_rows = []

        next_line = (
            SalesOrderLineItem.objects.filter(order=sales_order).aggregate(
                max_line=Max("line_int")
            )["max_line"]
            or 0
        ) + 1

        with transaction.atomic():
            for entry in entries:
                product_name = entry.get("product_name", "")
                quantity_raw = entry.get("quantity", None)
                row = entry.get("row", None)
                preview_row = {
                    "row": row,
                    "input": product_name,
                    "quantity": None,
                    "status": "skipped",
                    "reason": None,
                    "matched_ipn": None,
                    "matched_name": None,
                }

                if not product_name:
                    skipped_count += 1
                    unresolved.append({
                        "row": row,
                        "product_name": "",
                        "reason": "missing_product_name",
                        "candidates": [],
                    })
                    preview_row["reason"] = "missing_product_name"
                    preview_rows.append(preview_row)
                    continue

                if quantity_raw in [None, "", "None"]:
                    skipped_count += 1
                    unresolved.append({
                        "row": row,
                        "product_name": product_name,
                        "reason": "missing_quantity",
                        "candidates": [],
                    })
                    preview_row["reason"] = "missing_quantity"
                    preview_rows.append(preview_row)
                    continue

                try:
                    quantity = Decimal(str(quantity_raw))
                    preview_row["quantity"] = str(quantity)
                except (InvalidOperation, TypeError):
                    skipped_count += 1
                    unresolved.append({
                        "row": row,
                        "product_name": product_name,
                        "reason": "invalid_quantity",
                        "candidates": [],
                    })
                    preview_row["reason"] = "invalid_quantity"
                    preview_rows.append(preview_row)
                    continue

                if quantity <= 0:
                    skipped_count += 1
                    unresolved.append({
                        "row": row,
                        "product_name": product_name,
                        "reason": "non_positive_quantity",
                        "candidates": [],
                    })
                    preview_row["reason"] = "non_positive_quantity"
                    preview_rows.append(preview_row)
                    continue

                part, part_issue = self._find_part_for_name(product_name)

                if part_issue is not None:
                    skipped_count += 1
                    unresolved.append({
                        "row": row,
                        "product_name": product_name,
                        "reason": part_issue.get("reason"),
                        "candidates": part_issue.get("candidates", []),
                    })
                    preview_row["reason"] = part_issue.get("reason")
                    preview_rows.append(preview_row)
                    continue

                if part is None:
                    skipped_count += 1
                    unresolved.append({
                        "row": row,
                        "product_name": product_name,
                        "reason": "part_not_found",
                        "candidates": [],
                    })
                    preview_row["reason"] = "part_not_found"
                    preview_rows.append(preview_row)
                    continue

                preview_row["matched_ipn"] = part.IPN
                preview_row["matched_name"] = part.name

                line = SalesOrderLineItem(
                    order=sales_order,
                    part=part,
                    quantity=quantity,
                    line=str(next_line),
                    reference="Imported item",
                )

                try:
                    line.full_clean()
                    would_create_count += 1
                    preview_row["status"] = "ready" if dry_run else "imported"
                    preview_rows.append(preview_row)

                    if not dry_run:
                        line.save()
                        next_line += 1
                        created_count += 1
                except ValidationError as exc:
                    skipped_count += 1
                    errors.append({
                        "row": row,
                        "product_name": product_name,
                        "error": str(exc),
                    })
                    preview_row["status"] = "error"
                    preview_row["reason"] = str(exc)
                    preview_rows.append(preview_row)

            if dry_run:
                transaction.set_rollback(True)

        return JsonResponse(
            {
                "dry_run": dry_run,
                "created_count": created_count,
                "would_create_count": would_create_count,
                "skipped_count": skipped_count,
                "errors": errors,
                "unresolved": unresolved,
                "preview_rows": preview_rows,
            },
            status=200,
        )

    # Custom UI panels
    def get_ui_panels(self, request, context: dict, **kwargs):
        """Return a list of custom panels to be rendered in the InvenTree user interface."""

        panels = []

        target_model = context.get("target_model", None)
        target_id = context.get("target_id", None)

        # Only display this panel on sales order detail pages
        if target_model == "salesorder" and target_id is not None:
            panels.append({
                "key": "so-line-item-import-panel",
                "title": "Import Sales Order Lines",
                "description": "Upload an Excel file to create sales order line items",
                "icon": "ti:file-import:outline",
                "source": self.plugin_static_file(
                    "Panel.js:RenderSOLineItemImportPanel"
                ),
                "context": {
                    "import_url": reverse(f"plugin:{self.slug}:import-so-lines"),
                },
            })

        return panels
